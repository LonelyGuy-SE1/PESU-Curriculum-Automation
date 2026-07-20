import base64
import json
import logging

import sentry_sdk
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import Response, StreamingResponse
from postgrest.exceptions import APIError

from app.models.chat import ChatMessagePayload, ChatSessionPayload, ChatSessionTitlePayload
from app.services.agent_tools import call_tool, list_tool_schemas
from app.services.attachments import extract_text
from app.services.curriculum import load_document_draft, refined_course
from app.services.errors import database_http_exception
from app.services.openrouter import OpenRouterError, context_length, stream_chat
from app.supabase import first_row, supabase

router = APIRouter()
logger = logging.getLogger(__name__)


def load_chat_session(session_id: int) -> dict:
    try:
        row = first_row(supabase.table("chat_sessions").select("*").eq("id", session_id))
    except APIError as exc:
        raise database_http_exception(exc) from exc
    if not row:
        raise HTTPException(status_code=404, detail="Chat session not found")
    return row


def chat_messages(session_id: int) -> list[dict]:
    rows = supabase.table("chat_messages").select("*").eq("session_id", session_id).order("id").execute().data

    if not rows:
        return rows

    all_ids: list[int] = []
    for row in rows:
        for att in (row.get("metadata") or {}).get("attachments") or []:
            if isinstance(att, dict) and att.get("id"):
                all_ids.append(int(att["id"]))

    known_ids = set(all_ids)

    if not known_ids:
        orphan_rows = (
            supabase.table("chat_attachments")
            .select("id,filename,content_type,size_bytes,status,error")
            .eq("session_id", session_id)
            .is_("message_id", "null")
            .execute().data
        )
        orphans = [r for r in orphan_rows if r["id"] not in known_ids]

        if orphans:
            last_assistant = None
            for row in reversed(rows):
                if row.get("role") == "assistant":
                    last_assistant = row
                    break
            if last_assistant:
                meta = last_assistant.get("metadata") or {}
                existing = meta.get("attachments") or []
                for o in orphans:
                    existing.append({"id": o["id"]})
                    all_ids.append(o["id"])
                    known_ids.add(o["id"])
                meta["attachments"] = existing
                last_assistant["metadata"] = meta

    if not all_ids:
        return rows

    att_rows = (
        supabase.table("chat_attachments")
        .select("id,filename,content_type,size_bytes,status,error")
        .eq("session_id", session_id)
        .in_("id", all_ids)
        .execute()
        .data
    )
    att_map = {str(a["id"]): a for a in att_rows}

    for row in rows:
        meta = row.get("metadata") or {}
        enriched = []
        for att in meta.get("attachments") or []:
            if isinstance(att, dict) and att.get("id"):
                full = att_map.get(str(att["id"]), {})
                enriched.append({
                    "id": att["id"],
                    "name": full.get("filename") or "attachment",
                    "type": full.get("content_type") or "",
                    "size": full.get("size_bytes") or 0,
                    "status": full.get("status") or "",
                    "error": full.get("error") or "",
                })
            else:
                enriched.append(att)
        meta["attachments"] = enriched
        row["metadata"] = meta

    return rows


def insert_chat_message(session_id: int, role: str, content: str, metadata: dict | None = None) -> dict:
    return (
        supabase.table("chat_messages")
        .insert({"session_id": session_id, "role": role, "content": content, "metadata": metadata or {}})
        .execute()
        .data[0]
    )


def update_attachment_message(session_id: int, attachment_ids: list[int], message_id: int) -> None:
    if attachment_ids:
        supabase.table("chat_attachments").update({"message_id": message_id}).eq("session_id", session_id).in_("id", attachment_ids).execute()


def attachment_ids(metadata: dict | None) -> list[int]:
    ids = []
    for item in (metadata or {}).get("attachments") or []:
        if isinstance(item, dict) and item.get("id"):
            ids.append(int(item["id"]))
    return ids


def attachment_context(session_id: int, metadata: dict | None) -> str:
    ids = attachment_ids(metadata)
    if not ids:
        return ""
    rows = supabase.table("chat_attachments").select("filename,status,error,extracted_text").eq("session_id", session_id).in_("id", ids).execute().data
    blocks = []
    for row in rows:
        name = row.get("filename") or "attachment"
        status = row.get("status") or ""
        text = str(row.get("extracted_text") or "").strip()
        if text:
            blocks.append(f"Attachment: {name}\n{text}")
        else:
            error = row.get("error") or "No extracted text"
            blocks.append(f"Attachment: {name}\nStatus: {status}. {error}")
    return "\n\n".join(blocks)


def stable_context(value: dict) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2)


def chat_system_prompt(session: dict) -> str:
    session_id = session.get("id")
    context = ""
    if session.get("refined_id"):
        course = refined_course(int(session["refined_id"]))
        context = stable_context({"active_session_id": session_id, "active_refined_id": session["refined_id"], "active_course": course})
    elif session.get("document_draft_id"):
        context = stable_context({"active_session_id": session_id, **load_document_draft(int(session["document_draft_id"]))})
    else:
        context = stable_context({"active_session_id": session_id})
    return f"""You are the Syntagma live editor assistant.
Be concise, practical, and specific to the active curriculum data.
Keep conversations professional and friendly. Do not use em dashes in any output. Use standard hyphens or commas instead.
Never modify a course title, course code, or other identifying fields unless explicitly asked by the user. Use the exact name and code the user provides. Do not embellish or add words the user did not say.
When the user asks you to do something (create a course, generate a report, etc.), first send a brief text message acknowledging the request (1-2 sentences like "I'll create that course for you." or "Working on the report now.") before calling any tools. This lets the user know you received their request and are working on it.
Always respond by calling a tool -- never state limitations or guess. The available tools handle course data, fetching URLs, searching the web, generating spreadsheets, generating reports, and creating drafts.
Never silently fail. If a tool returns an error, report it to the user. If you finish a chain of tools, always call signal_done with a summary of what was accomplished.
When you need clarification before proceeding (credit category, semester, course code, elective vs core, specialization track, etc.), call ask_user with your question. The stream ends and the user will respond in their next message. Do NOT guess or assume when the user's intent is ambiguous.

Curriculum structure (B.Tech CSE):
- 8 semesters. Semesters 1-4: foundation and core courses (programming, math, basic sciences). Semesters 5-6: core courses plus elective specialization tracks. Semesters 7-8: advanced electives and capstone project.
- Credit categories: "5" = Core Course-Lab Integrated (4L 0T 2P 5C), "4" = Core Course (4L 0T 0P 4C), "2" = Core Theory (2L 0T 0P 2C), "0" = Foundation Course (0L 0T 0P 0C).
- Only credit_category "5" courses have a lab component (lab_experiments). All other courses (categories 4, 2, 0) must have empty lab_experiments.
- All target departments map to "B. TECH" program.
- Course code convention: prefix UE (university elective) or UZ (open elective), 2-digit batch year, 2-letter department (CS/MA/AM/etc), 3-digit level, letter suffix for semester parity (A=odd, B=even). Specialization electives have extended suffixes like AAX, BBX. Every course code must be unique across the entire curriculum.
- Semesters 5-6 have elective specialization tracks (e.g. Machine Intelligence and Data Science, Cybersecurity, etc.). Tracks are labeled with letters (A, B, C...). Each track has a set of elective courses assigned to it.
- Specialization tables (when listing courses under a specialization track) must only include elective courses (is_elective=true). Core courses and foundation courses must not appear in specialization track tables.
- Unit hours: every unit must be assigned exactly 14 hours. Do not vary hours across units.
- Deterministic/protected fields (require update_deterministic_fields to change): program, lecture_hours, tutorial_hours, practical_hours, self_study, credits, course_type. These are auto-computed from credit_category and target_department.
- Agent-editable fields: course_code, course_title, semester, is_elective, tools_languages, desirable_knowledge, prelude, objectives, course_outcomes, units, lab_experiments, text_books, reference_books.
- Fields that are lists (objectives, course_outcomes, text_books, reference_books, lab_experiments) are JSON arrays. Send them as ["item1", "item2"] in the tool call.
- desirable_knowledge should only reference courses that actually exist in the curriculum (use list_courses or get_course_codes to check). For brand-new courses with no prior dependencies, set it to an empty string. Never invent course names for desirable_knowledge.
- Courses have a status: "draft" (newly created, not yet finalized), "refined" (approved and visible in curriculum), "archived" (hidden).
- When a course is in "draft" status, you can update it directly with create_refined_course (pass the refined_id). Do not create a separate draft for draft-status courses.

Course types and electives:
- is_elective is a boolean flag separate from credit_category/course_type. A course with credit_category "4" can be either a core course (is_elective=false) or an elective course (is_elective=true).
- credit_category determines the hours and credits. is_elective determines whether the course appears in specialization track tables.
- To create an elective: set is_elective=true in create_refined_course. Use a course code with extended suffix (e.g. UE25CS342BA7 for semester 6).
- To assign an elective to specialization tracks: use assign_elective_to_tracks with the refined_id and specialization_ids.
- To create a new specialization track: use define_specialization.
- Electives can exist without being assigned to any track. Assignment to tracks is optional.

Syllabus design rules (enforce strictly when creating or modifying courses):
- Every course must have 3-5 units. Each unit has a title, content (syllabus text), and exactly 14 hours.
- Every course must have 3-4 course_outcomes that are measurable (use verbs like "Analyze", "Implement", "Design", "Evaluate").
- Every course must have 3-4 objectives describing what students will learn.
- text_books: at least 1 textbook reference with author, title, edition, publisher, year.
- reference_books: optional but recommended. Follow the same format as text_books.
- Course titles should be concise (under 80 chars) and descriptive.
- course_outcomes should NOT duplicate objectives. Outcomes are measurable assessments; objectives are learning goals.
- For credit_category "5": include 4-6 lab_experiments describing hands-on tasks.
- For credit_category "4" or "2": lab_experiments must be empty.
- prelude: 2-3 sentences describing the course scope and relevance.
- desirable_knowledge: list prerequisite courses or skills. Use only course codes that exist in the curriculum.

Decision guidelines:
- Before creating a new course, ALWAYS call list_courses to see all existing course codes. Generate a unique course code that does not appear in the list. The server also validates uniqueness and will reject duplicates.
- If the user asks to create something new, first check if it already exists using list_courses or get_course_codes. If it does NOT exist, proceed directly to create_refined_course with all required fields. Do NOT stop after just checking -- always complete the full workflow.
- If modifying an existing course that has status "refined", use create_course_draft (creates a reviewable draft).
- If modifying a course that has status "draft", use create_refined_course with the refined_id (direct update, no draft needed).
- When assigning electives to specializations, first list existing specializations with list_specializations, then use assign_elective_to_tracks.
- When unsure about any parameter (which semester, credit category, course code, elective vs core, etc.), call ask_user to ask the user. Do NOT guess or assume.
- When generating course content (units, objectives, etc.), follow the syllabus design rules above.

Chaining tools: When a user request clearly requires multiple steps (e.g. "export semester 3 to CSV" needs list_courses then batch_read_courses then create_spreadsheet), chain the tools in a single turn. Do not stop after one tool if the task is not yet complete. Stop chaining and respond only when the task is done or you need user input.
Always call signal_done at the end of a completed task with a concise summary.

Read source documents with get_attachment_text, then call create_report to save generated content as a chat attachment.

Tool selection for course changes (critical):
- create_course_draft: For modifying an existing course that has status "refined". Creates a reviewable draft the user must approve. Requires refined_id and a "fields" object containing ONLY the fields to change.
- update_agent_draft: For modifying a draft you already created (instead of creating a duplicate). Requires the draft_id from a previous create_course_draft call and a "fields" object with the changes. Use this when you need to adjust a draft after reviewing it.
- create_refined_course: For creating a brand-new course OR updating a course that still has status "draft". Pass all course details as flat arguments. For updates, include the refined_id.
- create_document_draft: For changes across multiple existing courses.
- update_deterministic_fields: ONLY for changing protected fields (program, hours, credits, course_type). Creates a blocked draft. Always confirm with the user first.

Before creating a course, always call list_courses or batch_read_courses to check if it already exists. If the course_code is not in the list, use create_refined_course. If it exists, use create_course_draft with the refined_id from the list.
When the user asks what changed, call diff_course_json or read the relevant draft before answering.
For broad document requests, use get_curriculum_json to inspect the whole syllabus before proposing edits. Prefer filtering by semester to avoid oversized responses.
For version comparison, call get_version to load a snapshot, then diff_versions to compare two versions.
For statistics and summaries, call get_curriculum_stats for aggregate data or batch_read_courses to read fields from many courses at once.
For spreadsheet exports, call batch_read_courses to gather data, then create_spreadsheet to generate CSV or Excel files.
For specialization management, call list_specializations to discover tracks, define_specialization to create one, and assign_elective_to_tracks / get_course_assignments to categorize electives.
To fetch a public URL, call fetch_url and use the returned text.
To search the web for current information, to ground a response in verified facts, or to compare with external documents, call web_search with a query. Always cite the source when using web search results.
Never apply a draft, never claim a draft was applied, and never claim the refined database was changed.
After creating a draft, tell the user to review the diff in the Review panel before applying it.
To change deterministic fields (program, hours, credits, course_type), call update_deterministic_fields. This creates a draft that is blocked until the user explicitly approves it in the Review panel. Confirm with the user before changing these fields.

Course data access -- prefer granular tools over full JSON:
- get_course_codes: lightweight IDs (refined_id, course_code, title, semester) -- use for lists/lookups
- get_course_syllabus: units, objectives, course_outcomes
- get_course_textbooks: text_books, reference_books
- get_course_deterministic: program, hours, credits, course_type (read-only, agent-protected)
- get_course_lab: lab_experiments, tools_languages
- get_course_fields: arbitrary specific fields (provide field name list)
- batch_read_courses: read specific fields from multiple courses in one call (preferred over looping get_course_fields)
- get_current_course_json: full course JSON -- only use when you truly need everything
- get_curriculum_json: full curriculum or filtered by semester (ALWAYS pass semester filter to avoid oversized responses)

When the user's request is fully addressed (draft created, question answered, report/spreadsheet generated), call signal_done with a concise summary of what was accomplished.

Active context:
{context or "No active course or document draft is selected."}"""


def model_messages(session_id: int, rows: list[dict]) -> tuple[list[dict], dict]:
    messages = []
    for row in rows:
        if row.get("role") not in {"user", "assistant"}:
            continue
        content = str(row.get("content") or "").strip()
        context = attachment_context(session_id, row.get("metadata")) if row.get("role") == "user" else ""
        if context:
            content = f"{content}\n\n{context}".strip()
        if content:
            messages.append({"role": row["role"], "content": content})

    budget = context_length() * 3
    total = sum(len(m["content"]) for m in messages)
    if total > budget and len(messages) >= 2:
        while total > budget and len(messages) >= 2:
            removed_user = messages.pop(0)
            total -= len(removed_user["content"])
            if messages:
                removed_assistant = messages.pop(0)
                total -= len(removed_assistant["content"])
        messages.insert(0, {
            "role": "user",
            "content": "[Earlier conversation context has been truncated to manage token usage]"
        })
    usage = {"total_chars": total, "budget_chars": budget, "message_count": len(messages), "context_length": context_length()}
    return messages, usage


def sse(event: str, data: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


@router.post("/chat/sessions")
def create_chat_session(payload: ChatSessionPayload):
    record = {"refined_id": payload.refined_id, "document_draft_id": payload.document_draft_id, "title": payload.title.strip()}
    result = supabase.table("chat_sessions").insert(record).execute()
    session = result.data[0]
    session_id = session["id"]

    if payload.refined_id:
        starter = f"👋 Hello! I'm ready to help with course **{payload.title}** (refined_id: {payload.refined_id}). You can ask me to:\n\n- Create a draft with specific changes\n- Explain the current syllabus structure\n- Compare with other courses\n- Fetch textbook details or lab experiments\n\nWhat would you like to do?"
    elif payload.document_draft_id:
        starter = f"👋 Hello! I'm ready to help with document draft **{payload.document_draft_id}**. You can ask me to:\n\n- Review the proposed changes across multiple courses\n- Compare draft versions\n- Generate a summary report\n\nWhat would you like to do?"
    else:
        starter = "👋 Hello! I'm your Syntagma assistant. You can ask me to:\n\n- Browse the curriculum with `get_curriculum_json`\n- Inspect specific courses with granular tools\n- Create drafts for course changes\n- Generate reports and comparisons\n\nWhat would you like to explore?"

    supabase.table("chat_messages").insert({
        "session_id": session_id,
        "role": "assistant",
        "content": starter,
        "metadata": {"starter": True}
    }).execute()

    return {"session": session}


@router.get("/chat/sessions")
def list_chat_sessions(refined_id: int | None = None, document_draft_id: int | None = None):
    query = supabase.table("chat_sessions").select("*").eq("status", "active")
    if refined_id is not None:
        query = query.eq("refined_id", refined_id)
    if document_draft_id is not None:
        query = query.eq("document_draft_id", document_draft_id)
    rows = query.order("id", desc=True).limit(50).execute().data
    return {"sessions": rows}


@router.get("/chat/sessions/{session_id}/messages")
def get_chat_messages(session_id: int):
    load_chat_session(session_id)
    return {"messages": chat_messages(session_id)}


@router.delete("/chat/sessions/{session_id}")
def clear_chat_session(session_id: int):
    load_chat_session(session_id)
    supabase.table("chat_sessions").delete().eq("id", session_id).execute()
    return {"message": "Chat deleted"}


@router.patch("/chat/sessions/{session_id}")
def rename_chat_session(session_id: int, payload: ChatSessionTitlePayload):
    load_chat_session(session_id)
    row = supabase.table("chat_sessions").update({"title": payload.title}).eq("id", session_id).execute().data[0]
    return {"message": "Chat renamed", "session": row}


@router.post("/chat/sessions/{session_id}/messages")
def create_chat_message(session_id: int, payload: ChatMessagePayload):
    if not payload.content and not payload.metadata:
        raise HTTPException(status_code=400, detail="Message content is required")
    session = load_chat_session(session_id)
    user_message = insert_chat_message(session_id, "user", payload.content, payload.metadata)
    update_attachment_message(session_id, attachment_ids(payload.metadata), user_message["id"])

    def stream():
        answer = []
        tool_results = []
        report_attachment_ids = []

        def remember_tool_result(name: str, result: dict) -> None:
            tool_results.append({"name": name, "result": result})
            attachment = (result or {}).get("attachment")
            if attachment and attachment.get("id"):
                report_attachment_ids.append(attachment["id"])

        def flush_tool_results():
            while tool_results:
                item = tool_results.pop(0)
                result = item["result"] or {}
                draft = result.get("draft")
                document_draft = result.get("document_draft")
                if item["name"] == "create_course_draft" and draft:
                    yield sse("draft", {"draft": draft})
                if item["name"] == "create_document_draft" and document_draft:
                    yield sse("document_draft", {"document_draft": document_draft})
                if item["name"] == "update_deterministic_fields" and draft:
                    yield sse("draft", {"draft": draft})
                if item["name"] == "create_refined_course" and result.get("refined_id"):
                    yield sse("refined_course", {"refined_id": result["refined_id"], "updated": result.get("updated", False)})

        def save_assistant_message():
            content = "".join(answer).strip()
            if not content:
                return None
            metadata = {}
            if report_attachment_ids:
                metadata["attachments"] = [{"id": aid} for aid in report_attachment_ids]
            message = insert_chat_message(session_id, "assistant", content, metadata)
            if report_attachment_ids:
                update_attachment_message(session_id, report_attachment_ids, message["id"])
            return message

        try:
            yield sse("status", {"message": "Analyzing your request..."})
            rows = chat_messages(session_id)
            system = chat_system_prompt(session)
            msgs, _ = model_messages(session_id, rows)
            for item in stream_chat(system, msgs, list_tool_schemas(), call_tool, remember_tool_result):
                if isinstance(item, dict) and "$status" in item:
                    yield sse("status", {"message": item["$status"]})
                    continue
                if isinstance(item, dict) and "$usage" in item:
                    yield sse("context_usage", item["$usage"])
                    continue
                if isinstance(item, dict) and "$event" in item:
                    event = item["$event"]
                    data = {k: v for k, v in item.items() if k != "$event"}
                    if event == "tool_call":
                        insert_chat_message(session_id, "tool", f"\u2699 {data.get('name', '')}({json.dumps(data.get('arguments', {}), ensure_ascii=False)})", {"name": data.get("name"), "arguments": data.get("arguments"), "tool_call_type": "call"})
                    elif event == "tool_result":
                        status_char = "\u2713" if data.get("status") == "ok" else "\u2717"
                        insert_chat_message(session_id, "tool", f"{status_char} {data.get('name', '')} completed", {"name": data.get("name"), "status": data.get("status"), "tool_call_type": "result"})
                    yield sse(event, data)
                    continue
                yield from flush_tool_results()
                answer.append(item)
                yield sse("token", {"text": item})
            yield from flush_tool_results()
            message = save_assistant_message()
            yield sse("done", {"message_id": message["id"] if message else 0})
        except OpenRouterError as exc:
            yield from flush_tool_results()
            save_assistant_message()
            logger.warning(
                "Chat model request failed for session %s: status=%s detail=%s",
                session_id,
                exc.status_code,
                exc.provider_message[:300],
            )
            err = exc.message
            if exc.provider_message:
                err = f"{err} ({exc.provider_message[:200]})"
            yield sse("error", {"message": err})
        except Exception as exc:
            yield from flush_tool_results()
            save_assistant_message()
            logger.exception("Chat stream failed for session %s", session_id)
            sentry_sdk.capture_exception(exc)
            yield sse("error", {"message": "An internal error occurred. Please try again later."})

    return StreamingResponse(stream(), media_type="text/event-stream", headers={"Cache-Control": "no-store"})


@router.post("/chat/sessions/{session_id}/attachments")
async def upload_chat_attachments(session_id: int, files: list[UploadFile] = File(...)):
    load_chat_session(session_id)
    attachments = []
    for file in files:
        data = await file.read()
        text, status, error = extract_text(file.filename or "attachment", file.content_type or "", data)
        
        # Store binary content as base64 for non-text files
        content_base64 = ""
        if file.content_type and not file.content_type.startswith("text/") and file.content_type != "application/json":
            content_base64 = base64.b64encode(data).decode()
        
        row = (
            supabase.table("chat_attachments")
            .insert(
                {
                    "session_id": session_id,
                    "filename": file.filename or "attachment",
                    "content_type": file.content_type or "",
                    "size_bytes": len(data),
                    "extracted_text": text,
                    "content_base64": content_base64,
                    "status": status,
                    "error": error,
                }
            )
            .execute()
            .data[0]
        )
        attachments.append(
            {
                "id": row["id"],
                "name": row["filename"],
                "type": row["content_type"],
                "size": row["size_bytes"],
                "status": row["status"],
                "error": row["error"],
                "extracted_chars": len(row.get("extracted_text") or ""),
            }
        )
    return {"attachments": attachments}


@router.get("/chat/sessions/{session_id}/attachments/{attachment_id}/download")
def download_chat_attachment(session_id: int, attachment_id: int):
    load_chat_session(session_id)
    row = supabase.table("chat_attachments").select("*").eq("id", attachment_id).eq("session_id", session_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found")
    row = row[0]
    
    content = row.get("content_base64")
    if content:
        data = base64.b64decode(content)
    else:
        data = (row.get("extracted_text") or "").encode()
    
    return Response(
        content=data,
        media_type=row.get("content_type") or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{row["filename"]}"', "Cache-Control": "no-store"},
    )


@router.get("/chat/sessions/{session_id}/attachments/{attachment_id}/preview")
def preview_chat_attachment(session_id: int, attachment_id: int):
    load_chat_session(session_id)
    row = supabase.table("chat_attachments").select("*").eq("id", attachment_id).eq("session_id", session_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found")
    row = row[0]

    content_type = row.get("content_type") or ""
    text = row.get("extracted_text") or ""
    b64 = row.get("content_base64") or ""
    filename = row.get("filename") or ""

    if content_type in ("text/markdown", "text/csv", "text/plain", "text/html") or content_type.startswith("text/"):
        return Response(
            content=text.encode(),
            media_type=content_type,
            headers={"Content-Disposition": "inline", "Cache-Control": "no-store"},
        )

    if b64:
        data = base64.b64decode(b64)
        if filename.lower().endswith(".xlsx") and data[:2] == b"PK":
            html = _xlsx_to_html(data)
            return Response(
                content=html.encode(),
                media_type="text/html",
                headers={"Content-Disposition": "inline", "Cache-Control": "no-store"},
            )
        return Response(
            content=data,
            media_type=content_type,
            headers={"Content-Disposition": "inline", "Cache-Control": "no-store"},
        )

    return Response(
        content=text.encode(),
        media_type="text/plain",
        headers={"Content-Disposition": "inline", "Cache-Control": "no-store"},
    )


def _xlsx_to_html(data: bytes) -> str:
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = ["<style>table.xlsx{border-collapse:collapse;font-family:system-ui;font-size:13px;width:100%}table.xlsx th{background:#00377b;color:#fff;padding:6px 10px;text-align:left;font-weight:600;border:1px solid #00377b}table.xlsx td{padding:5px 10px;border:1px solid #d1d5db}table.xlsx tr:nth-child(even){background:#f4f4f5}</style>"]
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        if len(wb.sheetnames) > 1:
            parts.append(f"<h4 style='margin:16px 0 8px;font-family:system-ui'>{ws.title}</h4>")
        parts.append("<table class='xlsx'><thead><tr>")
        for cell in rows[0]:
            parts.append(f"<th>{_esc(str(cell) if cell is not None else '')}</th>")
        parts.append("</tr></thead><tbody>")
        for row in rows[1:]:
            parts.append("<tr>")
            for cell in row:
                parts.append(f"<td>{_esc(str(cell) if cell is not None else '')}</td>")
            parts.append("</tr>")
        parts.append("</tbody></table>")
    wb.close()
    return "\n".join(parts)


def _esc(s: str) -> str:
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
